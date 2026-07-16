

import sys
import os
import msvcrt
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    print("  openpyxl install ho raha hai, ruko...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR   = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "Stefen_Scanned_Barcodes.xlsx"


def get_workbook():
    if EXCEL_FILE.exists():
        wb       = openpyxl.load_workbook(str(EXCEL_FILE))
        ws       = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                existing.add(str(row[1]).strip())
        return wb, ws, existing

    # Naya file banao
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scanned Barcodes"

    hfill = PatternFill("solid", start_color="1D6F42")
    hfont = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    hali  = Alignment(horizontal="center", vertical="center")
    ht    = Side(style="thin", color="999999")
    hbdr  = Border(left=ht, right=ht, top=ht, bottom=ht)

    for col, hdr in enumerate(["S.No.", "Barcode Number", "Date", "Time"], 1):
        c = ws.cell(1, col, hdr)
        c.font = hfont; c.fill = hfill
        c.alignment = hali; c.border = hbdr

    ws.row_dimensions[1].height = 26
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    wb.save(str(EXCEL_FILE))
    return wb, ws, set()


def save_barcode(barcode: str):
    barcode = barcode.strip()
    if not barcode:
        return "empty"

    wb, ws, existing = get_workbook()

    if barcode in existing:
        return "duplicate"

    thin = Side(style="thin", color="DDDDDD")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ali  = Alignment(horizontal="center", vertical="center")
    font = Font(name="Courier New", size=11)
    alt  = PatternFill("solid", start_color="F0FAF4")

    next_row = ws.max_row + 1
    now      = datetime.now()

    for col, val in enumerate([next_row-1, barcode, now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S")], 1):
        c = ws.cell(next_row, col, val)
        c.font = font; c.alignment = ali; c.border = bdr
        if next_row % 2 == 0:
            c.fill = alt

    wb.save(str(EXCEL_FILE))
    return "saved"


def main():
    os.system("cls")
    print("=" * 52)
    print("   STEFEN BARCODE SCANNER  →  EXCEL SAVER")
    print("=" * 52)
    print(f"   File: {EXCEL_FILE.name}")
    print(f"   Folder: {BASE_DIR}")
    print()
    print("   ✔  Scan karo → sequentially save hoga")
    print("   ✔  Duplicate scan → skip hoga")
    print("   ✔  Sab ek hi Excel file mein rehga")
    print()
    print("   Band karne ke liye: window band karo")
    print("=" * 52)
    print()
    print("   Ready — scanner se scan karo...\n")

    buf   = ""
    total = 0

    while True:
        try:
            ch = msvcrt.getwch()

            if ch in ('\r', '\n'):
                barcode = buf.strip()
                buf = ""
                if not barcode:
                    continue

                result = save_barcode(barcode)
                ts = datetime.now().strftime("%H:%M:%S")

                if result == "saved":
                    total += 1
                    print(f"   [{ts}]  ✓ SAVED      →  {barcode}   (Total: {total})")
                elif result == "duplicate":
                    print(f"   [{ts}]  ⚠ DUPLICATE  →  {barcode}   (skip)")

            elif ch == '\x03':
                raise KeyboardInterrupt
            else:
                buf += ch

        except KeyboardInterrupt:
            print()
            print(f"   Band ho gaya. Total saved: {total}")
            input("   Enter dabao band karne ke liye...")
            break
        except Exception:
            buf = ""


if __name__ == "__main__":
    main()
