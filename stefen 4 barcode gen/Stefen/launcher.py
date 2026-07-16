"""
Stefen Barcode Scanner Listener
================================
Background mein chalta hai.
Barcode scanner (USB/wireless) se scan karo
→ Automatically Excel mein save hota hai.

No window, no UI. Completely silent.
"""

import sys
import os
import json
import threading
import time
from datetime import datetime
from pathlib import Path

# ── Excel save path ──
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "Stefen_Scanned_Barcodes.xlsx"
LOG_FILE   = BASE_DIR / "scan_log.json"

# ── Load existing log ──
def load_log():
    if LOG_FILE.exists():
        try:
            return json.loads(LOG_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return []

def save_log(data):
    LOG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Save to Excel ──
def save_to_excel(entries):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return False

    thin  = Side(style="thin", color="DDDDDD")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_ali = Alignment(horizontal="center", vertical="center")
    mono  = Font(name="Courier New", size=11)
    alt   = PatternFill("solid", start_color="F0FAF4")

    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
        ws = wb.active
        next_row = ws.max_row + 1
        # Collect already saved barcodes
        saved = {str(ws.cell(r, 2).value) for r in range(2, ws.max_row+1) if ws.cell(r, 2).value}
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scanned Barcodes"
        # Header
        hfill = PatternFill("solid", start_color="1D6F42")
        hfont = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        for col, hdr in enumerate(["S.No.", "Barcode Number", "Scan Date", "Scan Time"], 1):
            c = ws.cell(1, col, hdr)
            c.font, c.fill, c.alignment, c.border = hfont, hfill, c_ali, bdr
        ws.row_dimensions[1].height = 26
        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        next_row = 2
        saved = set()

    added = 0
    for entry in entries:
        bc = str(entry["barcode"])
        if bc in saved:
            continue
        sno = next_row - 1
        row = [sno, bc, entry["date"], entry["time"]]
        for col, val in enumerate(row, 1):
            c = ws.cell(next_row, col, val)
            c.font      = mono
            c.alignment = c_ali
            c.border    = bdr
            if next_row % 2 == 0:
                c.fill = alt
        saved.add(bc)
        next_row += 1
        added += 1

    wb.save(str(EXCEL_FILE))
    return added


# ══════════════════════════════════════════
#   HTTP SERVER  (localhost:7734)
#   app.html sends scanned barcode here
# ══════════════════════════════════════════
import http.server

PORT = 7734

class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def log_message(self, *a): pass  # silent

    def do_GET(self):
        if self.path in ("/", ""):
            self.send_response(302)
            self.send_header("Location", "/app.html")
            self.end_headers()
        else:
            super().do_GET()

    def do_OPTIONS(self):
        self._cors(); self.end_headers()

    def _cors(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "POST,GET,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_POST(self):
        if self.path == "/scan":
            length = int(self.headers.get("Content-Length", 0))
            body   = json.loads(self.rfile.read(length))
            barcode = str(body.get("barcode", "")).strip()

            resp = {"ok": False}
            if barcode:
                now  = datetime.now()
                entry = {
                    "barcode": barcode,
                    "date":    now.strftime("%d/%m/%Y"),
                    "time":    now.strftime("%H:%M:%S"),
                }
                log = load_log()
                # Check duplicate
                existing = {e["barcode"] for e in log}
                if barcode not in existing:
                    log.append(entry)
                    save_log(log)
                    added = save_to_excel([entry])
                    resp = {"ok": True, "status": "saved", "added": added}
                else:
                    resp = {"ok": True, "status": "duplicate"}

            self._cors()
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(resp).encode())

        elif self.path == "/export":
            # Re-export all from log to Excel
            log = load_log()
            added = save_to_excel(log) if log else 0
            self._cors()
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            resp = {"ok": True, "file": "Stefen_Scanned_Barcodes.xlsx", "total": len(log)}
            self.wfile.write(json.dumps(resp).encode())

        elif self.path == "/stats":
            log = load_log()
            self._cors()
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            resp = {"ok": True, "total": len(log)}
            self.wfile.write(json.dumps(resp).encode())

        else:
            self.send_response(404); self.end_headers()


def start_server():
    with http.server.HTTPServer(("localhost", PORT), Handler) as s:
        s.serve_forever()


if __name__ == "__main__":
    import webbrowser

    # Start HTTP server in background thread
    t = threading.Thread(target=start_server, daemon=True)
    t.start()

    # Open browser
    time.sleep(0.8)
    webbrowser.open(f"http://localhost:{PORT}/app.html")

    # Keep main thread alive
    try:
        while True:
            time.sleep(60)
    except KeyboardInterrupt:
        pass
